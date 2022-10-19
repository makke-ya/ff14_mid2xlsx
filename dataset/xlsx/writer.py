# coding: utf-8
from copy import deepcopy
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Side, Border
from openpyxl.styles.alignment import Alignment
from more_itertools import chunked

from .base import XlsxIOBase
from ._static_data import (
    DICT_FOR_CONVERT_GERMAN2JAPAN,
    DICT_FOR_CONVERT_GERMAN2JAPAN_SHORTEN,
    DICT_FOR_OCTAVE
)


class XlsxWriter(XlsxIOBase):
    """
    EMS形式のXlsx書き込み用変数とヘルパー関数
    """
    def __init__(
        self,
        sheet_name="test",
        title="test",
        tempo=60,
        num_measures_in_system=4,
        start_measure_num=0,
        player_width=1.35,
        instrument_width=3.5,
        score_width=29.76,
        score_height=3.0,
        mark_width=0.7,
        shorten_width=0.5,
        header_width=2.5,
        title_width=8.0,
        color_width=1.35,
        shorten=False,
        rest_color="ff999999",
        octave0_color="ffc9daf8",
        octave1_color="fffff2cc",
        octave2_color="ffead1dc",
        octave3_color="ffff00ff",
        border_color="ff000000",
        non_border_color="ffc7c8c8",
    ):
        """
        Parameters
        ----------
        sheet_name : str, optional
            エクセルのシート名, by default "test"
        title : str, optional
            記載するタイトル名, by default "test"
        tempo : int, optional
            記載するテンポ, by default 60
        num_measures_in_system : int, optional
            1段に記譜する小節数, by default 4
        start_measure_num : int, optional
            何小節から開始するか, by default 0
        player_width : float, optional
            奏者欄のセル横幅, by default 1.35
        instrument_width : float, optional
            楽器欄のセル横幅, by default 3.5
        score_width : float, optional
            楽譜欄のセル横幅合計, by default 29.76
        score_height : float, optional
            楽譜の各セルの高さ, by default 1.5
        mark_width : float, optional
            楽譜行文字（[A],[B],[C],...)のセル横幅, by default 0.7
        shorten_width : float, optional
            短縮音表記にするセル幅, by default 0.5
        header_width : float, optional
            曲名、メトロノーム等のヘッダのセル横幅, by default 2.7
        title_width : float, optional
            タイトルのセル横幅, by default 8.0
        color_width : float, optional
            楽譜右上に描写する音の高さの最低セル横幅, by default 1.35
        shorten : bool, optional
            音名を半角にするかどうか, by default False
        rest_color : str, optional
            休符の色, by default "ff999999"
        octave0_color : str, optional
            オクターブ-1の色, by default "ffc9daf8"
        octave1_color : str, optional
            オクターブ0の色, by default "fffff2cc"
        octave2_color : str, optional
            オクターブ+1の色, by default "ffead1dc"
        octave3_color : str, optional
            オクターブ+2の色, by default "ffff00ff"
        border_color : str, optional
            楽譜の罫線の色, by default "ff000000"
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.sheet_name = sheet_name
        self.title = title
        self.tempo = tempo
        self.num_measures_in_system = num_measures_in_system
        self.start_measure_num = start_measure_num
        self.max_num_beats_in_row = 0
        self.player_width = player_width
        self.instrument_width = instrument_width
        self.score_width = score_width
        self.score_height = score_height
        self.mark_width = mark_width
        self.shorten_width = self._convert_cm2width(shorten_width)
        self.header_width = self._convert_cm2width(header_width)
        self.title_width = self._convert_cm2width(title_width)
        self.color_width = self._convert_cm2width(color_width)
        self.shorten = shorten
        self._initial_setting(self.ws)
        t = "solid"
        self.rest_color_fill = PatternFill(
            patternType=t, fgColor=rest_color, bgColor=rest_color
        )
        self.octave_color_fill = [
            PatternFill(patternType=t, fgColor=octave0_color, bgColor=octave0_color),
            PatternFill(patternType=t, fgColor=octave1_color, bgColor=octave1_color),
            PatternFill(patternType=t, fgColor=octave2_color, bgColor=octave2_color),
            PatternFill(patternType=t, fgColor=octave3_color, bgColor=octave3_color),
        ]
        self.thick_side = Side(color=border_color, style="thick")
        self.medium_side = Side(color=border_color, style="medium")
        self.thin_side = Side(color=border_color, style="thin")
        self.hair_side = Side(color=border_color, style="hair")
        self.dot_side = Side(color=border_color, style="dotted")
        self.non_border_side = Side(color=non_border_color, style="thin")
        self.note_font = Font(name="MS Pゴシック", sz=10, b=True)
        self.measure_font = Font(name="MS Pゴシック", sz=10, b=False)
        self.measure_align = Alignment(horizontal="left")
        self.mark_font = Font(name="Arial", sz=11, b=True)
        self.mark_align = Alignment(horizontal="center", vertical="center")
        self.header_font = Font(name="MS Pゴシック", sz=11, b=False)
        self.header_align = Alignment(horizontal="center", vertical="bottom")
        self.player_font = Font(name="MS Pゴシック", sz=11, b=False)
        self.player_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.instrument_font = Font(name="MS Pゴシック", sz=11, b=False)
        self.instrument_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def add_common_data_list(self, common_data_list, start_row=10, start_column=3, progress_bar=None):
        """
        共通音楽データ(CommonSoundData)のリストを記述する関数
        
        Parameters
        ----------
        common_data_list : list of CommonSoundData
            共通音楽データリスト
        start_row : int, optional
            楽譜記述を開始する行, by default 9
        start_column : int, optional
            楽譜記述を開始する列, by default 3
        progress_bar : tk.ProgressBar or None, optional
            プログレスバー, by default None
            Noneの場合は何もしない
        """
        note_lists = [common_data.get_note_list() for common_data in common_data_list]
        rate_lists = [common_data.get_rate_list() for common_data in common_data_list]
        for _ in range(self.start_measure_num):
            for note_list, rate_list in zip(note_lists, rate_lists):
                note_list.pop(0)
                rate_list.pop(0)

        # 行方向の最大ビート数をわり出しておく
        self._get_max_beats_in_row(rate_lists)

        # 1ビート内の必要なセル数、太字罫線を引く場所を得る
        num_cells_maps = self._get_required_num_cells(rate_lists)
        borders_maps = self._get_borders(note_lists)
        print("NUM_CELLS_MAPS", num_cells_maps)

        # cells_mapsを段ごとのセル数の最小公倍数に調整する
        num_cells_maps = self._adjust_cell_maps(num_cells_maps)

        # セル幅の調整
        self._adjust_cell_width(
            num_cells_maps[0][0], start_column=start_column + 1
        )  # マーカー分＋１

        # 曲名やメトロノームなど記載
        self._plot_header(num_cells_maps[0][0], start_column)

        progress_tick = 100 // len(common_data_list)
        _start_row = start_row
        for i, (common_data, note_list, num_cells_map, borders_map) in enumerate(zip(
            common_data_list, note_lists, num_cells_maps, borders_maps
        )):
            end_row = self.add_common_data(
                common_data,
                note_list,
                num_cells_map,
                borders_map,
                _start_row,
                start_column=start_column,
                progress_bar=progress_bar,
                progress_amount=progress_tick,
            )
            _start_row = end_row + 1
            
            # 進み具合を表示する場合は更新
            if progress_bar:
                progress_bar["value"] = progress_tick * (i + 1)
                progress_bar.update()

    def add_common_data(
        self, common_data, note_list, num_cells_map, borders_map,
        start_row, start_column=3, progress_bar=None, progress_amount=None
    ):
        print("note_list", note_list)
        print("num_cells_map", num_cells_map)
        print("borders_map", borders_map)
        now_row = start_row
        sum_cells = sum(num_cells_map[0])
        bef_sound = [None, None]
        mark_idx = 1
        now_measure = 1
        if progress_bar:
            progress_tick = (1.0 / (len(note_list) / self.num_measures_in_system)) * progress_amount
            progress_bar_value = progress_bar["value"]

        for system_idx, notes_in_system in enumerate(
            chunked(note_list, self.num_measures_in_system)
        ):
            print("system_idx", system_idx)
            num_cells_in_system = num_cells_map[system_idx]
            borders_in_system = borders_map[system_idx]
            beat_idx = 0
            now_column = start_column
            # 小節番号をかく
            now_measure = self._plot_measure_nums(
                now_row,
                now_column + 1,
                num_cells_in_system,
                borders_in_system,
                now_measure,
            )
            now_row += 1
            # マーカーを書く
            mark_idx = self._plot_marks(
                now_row, [now_column, now_column + sum_cells + 1], mark_idx
            )
            self._adjust_cell_height(now_row, now_row)
            now_column += 1
            for notes_in_measure in notes_in_system:
                merge_cell_nums = []
                bef_sound_in_measure = None
                bef_has_note = True
                for j, notes_in_beat in enumerate(notes_in_measure):
                    num_cells = num_cells_in_system[beat_idx]
                    borders = borders_in_system[beat_idx]
                    print("row, column", now_row, now_column)
                    print("yes", notes_in_beat, num_cells, borders)
                    cells_per_notes = num_cells // len(notes_in_beat)
                    self._plot_beat_border(
                        now_row,
                        now_column,
                        now_row,
                        now_column + num_cells - 1,
                        borders,
                    )
                    has_note = False
                    for i, notes in enumerate(notes_in_beat):
                        if i == 0 and j != 0 and j != len(notes_in_measure) // 2:
                            continue
                        if type(notes) != str:
                            has_note = True
                            break
                    for notes in notes_in_beat:
                        if has_note is True and type(notes) != str and len(notes) != 0:
                            merge_cell_nums.append(
                                [(now_row, now_column), (now_row, now_column)]
                            )
                            bef_sound_in_measure = notes
                        elif (
                            has_note is True
                            and bef_has_note is True
                            and notes == "-"
                            and bef_sound_in_measure is not None
                        ):
                            merge_cell_nums[-1][1] = (
                                now_row,
                                now_column + cells_per_notes - 1,
                            )
                        elif notes == "r":
                            bef_sound_in_measure = None

                        # 結合セル
                        self._merge_cells(
                            now_row,
                            now_column,
                            now_row,
                            now_column + cells_per_notes - 1,
                        )
                        width_dict = self._get_cell_widths(
                            self.ws, now_column, now_column + cells_per_notes - 1
                        )
                        # shorten = (
                        #     True
                        #     if sum(width_dict.values()) < self.shorten_width
                        #     else False
                        # )
                        shorten = self.shorten
                        print("width", width_dict)
                        # 音をプロット
                        bef_sound = self._plot_sound(
                            now_row, now_column, notes, bef_sound, shorten
                        )
                        now_column += cells_per_notes
                    bef_has_note = has_note
                    beat_idx += 1
                # 伸ばすセルは結合（拍を超えていても結合、小節を挟んだ場合は結合しない）
                if len(merge_cell_nums) > 1:
                    for (_row1, _col1), (_row2, _col2) in merge_cell_nums:
                        self._merge_cells(_row1, _col1, _row2, _col2)
            now_row += 1

            if progress_bar:
                progress_bar_value += progress_tick
                progress_bar["value"] = int(progress_bar_value)
                progress_bar.update()

        # 奏者、楽器を書く
        self._plot_player_and_instrument(
            start_row,
            now_row,
            start_column,
            player_idx=common_data.get_player_idx(),
            instrument_name=common_data.get_program_str(),
            difficulty=common_data.get_difficulty(self.tempo)
        )
        return now_row

    def _initial_setting(self, sheet):
        sheet.column_dimensions["A"].width = self._convert_cm2width(self.player_width)
        sheet.column_dimensions["B"].width = self._convert_cm2width(
            self.instrument_width
        )

    def _plot_header(self, num_cells, start_column=3):
        today_str = date.today()
        now_row = 1
        now_col = start_column
        # 改定日
        self._plot_cell(now_row, now_col, val="改定日：{}".format(today_str))
        now_row += 1
        now_col += 1

        # セルの幅を取得
        sum_cells = sum(num_cells)
        width_list = sorted(
            self._get_cell_widths(self.ws, now_col, now_col + sum_cells - 1).items(),
            key=lambda x: x[0],
        )
        print("width_list:", width_list)

        # 曲名
        border = Border(
            left=self.medium_side, right=self.medium_side,
            top=self.medium_side, bottom=self.medium_side,
        )
        idx = 0
        idx = self._get_next_width_idx(width_list, idx, self.header_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, val="曲名", border=border)
        now_col = width_list[idx][0]

        # タイトル
        idx = self._get_next_width_idx(width_list, idx, self.title_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, val=self.title, border=border)
        self._plot_cell(now_row, width_list[idx - 1][0], border=border)
        now_col = width_list[idx][0]

        # メトロノーム
        idx = self._get_next_width_idx(width_list, idx, self.header_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, val="メトロノーム", border=border)
        now_col = width_list[idx][0]

        # メトロノームの値
        idx = self._get_next_width_idx(width_list, idx, self.header_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, val=str(self.tempo), border=border)
        self._plot_cell(now_row, width_list[idx - 1][0], border=border)
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        print("WIDTH_LIST", idx, width_list)
        now_col = width_list[idx][0]
        color_start_idx = idx

        # 低音
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row - 1, now_col, val="低音bass")
        self._plot_cell(now_row, now_col, fill=self.octave_color_fill[0])
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        if len(width_list) <= idx:
            idx = color_start_idx
            now_row += 2
        now_col = width_list[idx][0]

        # 中音
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row - 1, now_col, val="中音mediant")
        self._plot_cell(now_row, now_col, fill=self.octave_color_fill[1])
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        if len(width_list) <= idx:
            idx = color_start_idx
            now_row += 2
        now_col = width_list[idx][0]

        # 高音
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row - 1, now_col, val="高音soprano")
        self._plot_cell(now_row, now_col, fill=self.octave_color_fill[2])
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        if len(width_list) <= idx:
            idx = color_start_idx
            now_row += 2
        now_col = width_list[idx][0]

        # 最高音
        idx = self._get_next_width_idx(width_list, idx, self.color_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._plot_cell(now_row - 1, now_col, val="最高音highest")
        self._plot_cell(now_row, now_col, fill=self.octave_color_fill[3])
        if len(width_list) <= idx:
            idx = color_start_idx
            now_row += 2
        now_col = width_list[idx][0]

        # 短縮版の音説明
        self._plot_cell(now_row, now_col, val="C=ド,F=ファ")

        # 動画&MIDI&編曲者
        idx = 0
        now_col = start_column + 1
        now_row = 3
        idx = self._get_next_width_idx(width_list, idx, self.header_width)
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._merge_cells(now_row + 1, now_col, now_row + 1, width_list[idx - 1][0])
        self._merge_cells(now_row + 2, now_col, now_row + 2, width_list[idx - 1][0])
        self._merge_cells(now_row + 3, now_col, now_row + 3, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, val="原曲URL", border=border)
        self._plot_cell(now_row + 1, now_col, val="FF演奏動画URL", border=border)
        self._plot_cell(now_row + 2, now_col, val="参考音源リンク", border=border)
        self._plot_cell(now_row + 3, now_col, val="編曲者", border=border)
        now_col = width_list[idx][0]

        # 動画&MIDI&編曲者欄
        idx = self._get_next_width_idx(
            width_list, idx, self.title_width + (self.header_width * 2)
        )
        self._merge_cells(now_row, now_col, now_row, width_list[idx - 1][0])
        self._merge_cells(now_row + 1, now_col, now_row + 1, width_list[idx - 1][0])
        self._merge_cells(now_row + 2, now_col, now_row + 2, width_list[idx - 1][0])
        self._merge_cells(now_row + 3, now_col, now_row + 3, width_list[idx - 1][0])
        self._plot_cell(now_row, now_col, border=border)
        self._plot_cell(now_row, now_col, border=border)
        self._plot_cell(now_row, width_list[idx - 1][0], border=border)
        self._plot_cell(now_row + 1, now_col, border=border)
        self._plot_cell(now_row + 1, width_list[idx - 1][0], border=border)
        self._plot_cell(now_row + 2, now_col, border=border)
        self._plot_cell(now_row + 2, width_list[idx - 1][0], border=border)
        self._plot_cell(now_row + 3, now_col, border=border)
        self._plot_cell(now_row + 3, width_list[idx - 1][0], border=border)

    def _get_next_width_idx(self, width_list, idx, limit):
        sum_width = 0
        for i, (_, width) in enumerate(width_list):
            if i < idx:
                continue
            sum_width += width
            if sum_width >= limit:
                idx = i + 1
                break
        return idx

    def _plot_sound(self, row, column, notes, bef_sound, shorten=False):
        if notes == "r":
            self._plot_cell(row, column, fill=self.rest_color_fill)
        elif notes == "-":
            if bef_sound[1] is None:
                self._plot_cell(row, column, fill=self.rest_color_fill)
            else:
                self._plot_cell(row, column, fill=self.octave_color_fill[bef_sound[1]])
        else:
            notes_str = ""
            fill = None
            # shorten = False if len(notes) < 2 else True
            for note in notes:
                split_note = note.split("_")
                octave = 1 if len(split_note) == 1 else DICT_FOR_OCTAVE[split_note[1]]
                fill = self.octave_color_fill[octave]
                if shorten is False:
                    scale = DICT_FOR_CONVERT_GERMAN2JAPAN[split_note[0]]
                else:
                    scale = DICT_FOR_CONVERT_GERMAN2JAPAN_SHORTEN[split_note[0]]
                notes_str += scale
                bef_sound = [scale, octave]
            self._plot_cell(row, column, notes_str, self.note_font, fill=fill)
        return bef_sound

    def _plot_marks(self, row, columns, mark_idx):
        for column in columns:
            column_str = self._convert_column_num(column)
            self.ws.column_dimensions[column_str].width = self._convert_cm2width(
                self.mark_width
            )
            val = self._convert_column_num(mark_idx).lower()
            border = Border(
                top=self.medium_side,
                left=self.medium_side,
                bottom=self.medium_side,
                right=self.medium_side,
            )
            self._plot_cell(row, column, val, self.mark_font, self.mark_align, border)
        return mark_idx + 1

    def _plot_beat_border(
        self, start_row, start_column, end_row, end_column, thick_flag
    ):
        start_cell_str = self._convert_cell_num(start_row, start_column)
        end_cell_str = self._convert_cell_num(end_row, end_column)
        print(start_row, start_column, "{}:{}".format(start_cell_str, end_cell_str))
        for row_cells in self.ws["{}:{}".format(start_cell_str, end_cell_str)]:
            for idx, cell in enumerate(row_cells):
                border = Border(
                    top=self.medium_side,
                    left=self.non_border_side,
                    bottom=self.medium_side,
                    right=self.non_border_side,
                )
                if idx == 0 and thick_flag[0] == 1:
                    border.left = self.thin_side
                elif idx == 0 and thick_flag[0] == 2:
                    border.left = self.medium_side
                if idx == len(row_cells) - 1 and thick_flag[1] == 1:
                    border.right = self.thin_side
                elif idx == len(row_cells) - 1 and thick_flag[1] == 2:
                    border.right = self.medium_side
                cell.border = border

    def _plot_measure_nums(
        self, row, column, num_cells_in_system, borders_in_system, start_measure
    ):
        merge_cell_num = 0
        num_measure = start_measure
        now_row = row
        now_column = column
        for num_cells, borders in zip(num_cells_in_system, borders_in_system):
            if merge_cell_num != 0 and borders[0] == 2:
                self._plot_measure_num(now_row, now_column, num_measure, merge_cell_num)
                num_measure += 1
                now_column += merge_cell_num
                merge_cell_num = 0
            merge_cell_num += num_cells
        if merge_cell_num != 0:
            self._plot_measure_num(now_row, now_column, num_measure, merge_cell_num)
            num_measure += 1
        return num_measure

    def _plot_measure_num(self, now_row, now_column, num_measure, merge_cell_num):
        print(now_row, now_column, merge_cell_num, num_measure)
        self._merge_cells(now_row, now_column, now_row, now_column + merge_cell_num - 1)
        self._plot_cell(
            now_row, now_column, num_measure, self.measure_font, self.measure_align
        )

    def _plot_player_and_instrument(
        self, start_row, end_row, start_column, player_idx, instrument_name,
        difficulty=None
    ):
        border = Border(
            top=self.medium_side,
            left=self.medium_side,
            bottom=self.medium_side,
            right=self.medium_side,
        )
        headers = ["楽器", "奏者"]
        player_str = (
            "奏者{}".format(player_idx) if difficulty is None
            else "奏者{}\n☆{}".format(player_idx, difficulty)
        )
        names = [instrument_name, player_str]
        fonts = [self.instrument_font, self.player_font]
        aligns = [self.instrument_align, self.player_align]
        print(start_row, end_row, start_column, player_idx, instrument_name)
        for i, (header, name, font, align) in enumerate(
            zip(headers, names, fonts, aligns), start=1
        ):
            col = start_column - i
            if col > 0:
                self._plot_cell(
                    start_row, col, header, self.header_font, self.header_align, border
                )
                self._merge_cells(start_row + 1, col, end_row - 1, col)
                self._plot_cell(start_row + 1, col, name, font, align, border)
                self._plot_cell(end_row - 1, col, border=border)

    def _plot_cell(
        self, row, col, val=None, font=None, align=None, border=None, fill=None
    ):
        cell_str = self._convert_cell_num(row, col)
        if val is not None:
            self.ws[cell_str].value = val
        if font is not None:
            self.ws[cell_str].font = font
        if align is not None:
            self.ws[cell_str].alignment = align
        if border is not None:
            self.ws[cell_str].border = border
        if fill is not None:
            self.ws[cell_str].fill = fill

    def _merge_cells(self, start_row, start_column, end_row, end_column):
        start_cell_str = self._convert_cell_num(start_row, start_column)
        end_cell_str = self._convert_cell_num(end_row, end_column)
        print("merge", start_cell_str, ":", end_cell_str)
        self.ws.merge_cells("{}:{}".format(start_cell_str, end_cell_str))

    def _get_max_beats_in_row(self, rate_lists):
        # 行方向の最大ビート数を得る
        for rate_list in rate_lists:
            for chunked_rates in chunked(rate_list, self.num_measures_in_system):
                temp_x = 0
                for rate_in_measure, _ in chunked_rates:
                    temp_x += len(rate_in_measure)
                if temp_x > self.max_num_beats_in_row:
                    self.max_num_beats_in_row = temp_x

    def _get_required_num_cells(self, rate_lists):
        num_cells_maps = []

        print(rate_lists)
        for rate_list in rate_lists:
            num_cells_map = []
            for chunked_rates in chunked(rate_list, self.num_measures_in_system):
                now_beat = 0
                num_cells_in_beat = [0 for x in range(self.max_num_beats_in_row)]
                for _, rate_in_beats in chunked_rates:
                    print(rate_in_beats)
                    for rates in rate_in_beats:
                        num_cells_in_beat[now_beat] = len(rates)
                        now_beat += 1
                num_cells_map.append(num_cells_in_beat)
            num_cells_maps.append(num_cells_map)

        print(num_cells_maps)
        return num_cells_maps

    def _get_borders(self, note_lists):
        borders_maps = []

        print(note_lists)
        for note_list in note_lists:
            borders_map = []
            for notes_in_system in chunked(note_list, self.num_measures_in_system):
                now_beat = 0
                border_in_beat = [[0, 0] for x in range(self.max_num_beats_in_row)]
                for notes_in_measure in notes_in_system:
                    for num_beat, notes_in_beat in enumerate(notes_in_measure):
                        if num_beat == 0:
                            border_in_beat[now_beat][0] = 2
                        else:
                            # elif notes_in_beat[0] != "-":
                            border_in_beat[now_beat][0] = 1
                        now_beat += 1
                border_in_beat[now_beat - 1][1] = 2
                borders_map.append(border_in_beat)
            borders_maps.append(borders_map)

        print(borders_maps)
        return borders_maps

    def _adjust_cell_maps(self, num_cells_maps):
        for now_beat in range(self.max_num_beats_in_row):
            num_list = [
                num_cells_in_beat[now_beat]
                for num_cells_map in num_cells_maps
                for num_cells_in_beat in num_cells_map
            ]
            if sum(num_list) == 0:
                continue
            num_lcm = self.calc_LCM(num_list)
            for map_idx, num_cells_map in enumerate(num_cells_maps):
                for system_idx, _ in enumerate(num_cells_map):
                    num_cells_maps[map_idx][system_idx][now_beat] = num_lcm
        return num_cells_maps

    def _adjust_cell_width(self, num_cells_in_system, start_column=4):
        beat_width = self.score_width / self.max_num_beats_in_row
        sheet = self.ws
        now_column = start_column
        for num_cells in num_cells_in_system:
            cell_width = beat_width / num_cells
            for _ in range(num_cells):
                column_str = self._convert_column_num(now_column)
                sheet.column_dimensions[column_str].width = self._convert_cm2width(
                    cell_width
                )
                now_column += 1

    def _adjust_cell_height(self, start_row, end_row):
        sheet = self.ws
        for row in range(start_row, end_row + 1):
            sheet.row_dimensions[row].height = self._convert_cm2width(
                self.score_height
            )

    def fwrite(self, filename):
        self.wb.save(filename)

    def calc_LCM(self, num_list):
        """
        再帰によってリスト内の数値全てにおける最小公倍数を求める関数
        
        Parameters
        ----------
        num_list : list of int
            数値リスト
        
        Returns
        -------
        int
            最小公倍数
        """
        _list = sorted(deepcopy(num_list))
        if len(_list) != 2:
            biggest_num = _list.pop()
            return lcm(biggest_num, self.calc_LCM(_list))
        else:
            return lcm(_list[1], _list[0])

    def calc_GCD(self, num_list):
        """
        再帰によってリスト内の数値全てにおける最大公約数を求める関数
        
        Parameters
        ----------
        num_list : list of int
            数値リスト
        
        Returns
        -------
        int
            最大公約数
        """
        _list = sorted(deepcopy(num_list))
        if len(_list) != 2:
            biggest_num = _list.pop()
            return gcd(biggest_num, self.calc_GCD(_list))
        else:
            return gcd(_list[1], _list[0])


class ThreeLineXlsxWriter(XlsxWriter):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.score_height = 2.2

    def add_common_data(
        self, common_data, note_list, num_cells_map, borders_map,
        start_row, start_column=3, progress_bar=None, progress_amount=None
    ):
        print("note_list", note_list)
        print("num_cells_map", num_cells_map)
        print("borders_map", borders_map)
        now_row = start_row
        sum_cells = sum(num_cells_map[0])
        bef_sound = [None, None]
        mark_idx = 1
        now_measure = 1
        if progress_bar:
            progress_tick = (1.0 / (len(note_list) / self.num_measures_in_system)) * progress_amount
            progress_bar_value = progress_bar["value"]

        for system_idx, notes_in_system in enumerate(
            chunked(note_list, self.num_measures_in_system)
        ):
            print("system_idx", system_idx)
            num_cells_in_system = num_cells_map[system_idx]
            borders_in_system = borders_map[system_idx]
            beat_idx = 0
            now_column = start_column
            # 小節番号をかく
            now_measure = self._plot_measure_nums(
                now_row,
                now_column + 1,
                num_cells_in_system,
                borders_in_system,
                now_measure,
            )
            now_row += 1
            # マーカーを書く
            mark_idx = self._plot_marks(
                [now_row, now_row + 2],
                [now_column, now_column + sum_cells + 1],
                mark_idx
            )
            self._adjust_cell_height(now_row, now_row + 2)
            now_column += 1
            for notes_in_measure in notes_in_system:
                merge_cell_nums = []
                bef_sound_in_measure = None
                bef_has_note = True
                for j, notes_in_beat in enumerate(notes_in_measure):
                    num_cells = num_cells_in_system[beat_idx]
                    borders = borders_in_system[beat_idx]
                    print("row, column", now_row, now_column)
                    print("yes", notes_in_beat, num_cells, borders)
                    cells_per_notes = num_cells // len(notes_in_beat)
                    self._plot_beat_border(
                        now_row,
                        now_column,
                        now_row + 2,
                        now_column + num_cells - 1,
                        borders,
                    )
                    has_note = False
                    for i, notes in enumerate(notes_in_beat):
                        if i == 0 and j != 0 and j != len(notes_in_measure) // 2:
                            continue
                        if type(notes) != str:
                            has_note = True
                            break
                    for notes in notes_in_beat:
                        if has_note is True and type(notes) != str and len(notes) != 0:
                            merge_cell_nums.append(
                                [(now_row, now_column), (now_row + 2, now_column)]
                            )
                            bef_sound_in_measure = notes
                        elif (
                            has_note is True
                            and bef_has_note is True
                            and notes == "-"
                            and bef_sound_in_measure is not None
                        ):
                            merge_cell_nums[-1][1] = (
                                now_row + 2,
                                now_column + cells_per_notes - 1,
                            )
                        elif notes == "r":
                            bef_sound_in_measure = None

                        # 結合セル
                        for _row in range(now_row, now_row + 3):
                            self._merge_cells(
                                _row,
                                now_column,
                                _row,
                                now_column + cells_per_notes - 1,
                            )
                        width_dict = self._get_cell_widths(
                            self.ws, now_column, now_column + cells_per_notes - 1
                        )
                        shorten = self.shorten
                        # shorten = (
                        #     True
                        #     if sum(width_dict.values()) < self.shorten_width
                        #     else False
                        # )
                        print("width", width_dict)
                        # 音をプロット
                        bef_sound = self._plot_sound(
                            now_row, now_column, notes, bef_sound, shorten
                        )
                        now_column += cells_per_notes
                    bef_has_note = has_note
                    beat_idx += 1
                # 伸ばすセルは結合（拍を超えていても結合、小節を挟んだ場合は結合しない）
                if len(merge_cell_nums) > 1:
                    for (_row1, _col1), (_row2, _col2) in merge_cell_nums:
                        for _row in range(_row1, _row2 + 1):
                            self._merge_cells(_row, _col1, _row, _col2)
            now_row += 3

            if progress_bar:
                progress_bar_value += progress_tick
                progress_bar["value"] = int(progress_bar_value)
                progress_bar.update()

        # 奏者、楽器を書く
        self._plot_player_and_instrument(
            start_row,
            now_row,
            start_column,
            player_idx=common_data.get_player_idx(),
            instrument_name=common_data.get_program_str(),
            difficulty=common_data.get_difficulty(self.tempo)
        )
        return now_row

    def _plot_marks(self, rows, columns, mark_idx):
        """
        楽譜の段のマーカーを描写する関数
        
        Parameters
        ----------
        rows : [type]
            [description]
        columns : [type]
            [description]
        mark_idx : [type]
            [description]
        
        Returns
        -------
        [type]
            [description]
        """
        for column in columns:
            for row in range(rows[0], rows[1] + 1):
                column_str = self._convert_column_num(column)
                self.ws.column_dimensions[column_str].width = self._convert_cm2width(
                    self.mark_width
                )
                val = self._convert_column_num(mark_idx).lower()
                border = Border(
                    top=self.medium_side,
                    left=self.medium_side,
                    bottom=self.medium_side,
                    right=self.medium_side,
                )
                self._plot_cell(row, column, val, self.mark_font, self.mark_align, border)
            self._merge_cells(rows[0], column, rows[1], column)
        return mark_idx + 1

    def _plot_beat_border(
        self, start_row, start_column, end_row, end_column, thick_flag
    ):
        for row in range(start_row, end_row + 1):
            start_cell_str = self._convert_cell_num(row, start_column)
            end_cell_str = self._convert_cell_num(row, end_column)
            print(start_row, start_column, "{}:{}".format(start_cell_str, end_cell_str))
            for row_cells in self.ws["{}:{}".format(start_cell_str, end_cell_str)]:
                for idx, cell in enumerate(row_cells):
                    border = Border(
                        top=self.non_border_side,
                        left=self.non_border_side,
                        bottom=self.non_border_side,
                        right=self.non_border_side,
                    )
                    if row == start_row:
                        border.top = self.medium_side
                        border.bottom = self.thin_side
                    if row == end_row:
                        border.top = self.thin_side
                        border.bottom = self.medium_side
                    if idx == 0 and thick_flag[0] == 1:
                        border.left = self.thin_side
                    elif idx == 0 and thick_flag[0] == 2:
                        border.left = self.medium_side
                    if idx == len(row_cells) - 1 and thick_flag[1] == 1:
                        border.right = self.thin_side
                    elif idx == len(row_cells) - 1 and thick_flag[1] == 2:
                        border.right = self.medium_side
                    cell.border = border

    def _plot_sound(self, row, column, notes, bef_sound, shorten=False):
        # まずは休符を書く
        for _row in range(row, row + 3):
            self._plot_cell(_row, column, fill=self.rest_color_fill)

        if notes == "r":
            bef_sound = [None, None]
        elif notes == "-":
            if bef_sound[1] is not None:
                margin = -1 * (bef_sound[1] - 2)  # 0 -> 2, 1 -> 1, 2,3 -> 0
                margin = margin if margin >= 0 else 0
                octave = 2 if bef_sound[1] == 3 else bef_sound[1]
                self._plot_cell(
                    row + margin, column, fill=self.octave_color_fill[octave]
                )
        else:
            for now_octave, margin in zip([0, 1, 2], [2, 1, 0]):
                notes_str = ""
                fill = None
                octave_notes = []
                for note in notes:
                    split_note = note.split("_")
                    octave = 1 if len(split_note) == 1 else DICT_FOR_OCTAVE[split_note[1]]
                    octave = 2 if octave == 3 else octave
                    if octave != now_octave:
                        continue
                    octave_notes.append(note)
                if len(octave_notes) == 0:
                    continue
                shorten = True if len(octave_notes) > 1 else shorten
                for note in octave_notes:
                    split_note = note.split("_")
                    octave = 1 if len(split_note) == 1 else DICT_FOR_OCTAVE[split_note[1]]
                    fill = self.octave_color_fill[now_octave]
                    if shorten is False:
                        scale = DICT_FOR_CONVERT_GERMAN2JAPAN[split_note[0]]
                    else:
                        scale = DICT_FOR_CONVERT_GERMAN2JAPAN_SHORTEN[split_note[0]]
                    if octave == 3:
                        scale += "^"
                    notes_str += scale
                    bef_sound = [scale, octave]
                self._plot_cell(
                    row + margin, column, notes_str, self.note_font, fill=fill)
        return bef_sound


class FlexibleLineXlsxWriter(ThreeLineXlsxWriter):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.score_height = 3.0

    def add_common_data(
        self, common_data, note_list, num_cells_map, borders_map,
        start_row, start_column=3, progress_bar=None, progress_amount=None
    ):
        print("note_list", note_list)
        print("num_cells_map", num_cells_map)
        print("borders_map", borders_map)
        now_row = start_row
        sum_cells = sum(num_cells_map[0])
        bef_sound = [None, None]
        mark_idx = 1
        now_measure = 1
        if progress_bar:
            progress_tick = (1.0 / (len(note_list) / self.num_measures_in_system)) * progress_amount
            progress_bar_value = progress_bar["value"]

        for system_idx, notes_in_system in enumerate(
            chunked(note_list, self.num_measures_in_system)
        ):
            num_rows = max([1] + [
                len(notes)
                for notes_in_measure in notes_in_system
                for notes_in_beat in notes_in_measure
                for notes in notes_in_beat if type(notes) is list
            ])
            print("system_idx", system_idx)
            print("num_rows", num_rows)
            num_cells_in_system = num_cells_map[system_idx]
            borders_in_system = borders_map[system_idx]
            beat_idx = 0
            now_column = start_column
            # 小節番号をかく
            now_measure = self._plot_measure_nums(
                now_row,
                now_column + 1,
                num_cells_in_system,
                borders_in_system,
                now_measure,
            )
            now_row += 1
            # マーカーを書く
            mark_idx = self._plot_marks(
                [now_row, now_row + num_rows - 1],
                [now_column, now_column + sum_cells + 1],
                mark_idx
            )
            self._adjust_cell_height(now_row, now_row + num_rows)
            now_column += 1
            for notes_in_measure in notes_in_system:
                merge_cell_nums = []
                bef_sound_in_measure = None
                bef_has_note = True
                for j, notes_in_beat in enumerate(notes_in_measure):
                    num_cells = num_cells_in_system[beat_idx]
                    borders = borders_in_system[beat_idx]
                    print("row, column", now_row, now_column)
                    print("yes", notes_in_beat, num_cells, borders)
                    cells_per_notes = num_cells // len(notes_in_beat)
                    self._plot_beat_border(
                        now_row,
                        now_column,
                        now_row + num_rows - 1,
                        now_column + num_cells - 1,
                        borders,
                    )
                    has_note = False
                    for i, notes in enumerate(notes_in_beat):
                        if i == 0 and j != 0 and j != len(notes_in_measure) // 2:
                            continue
                        if type(notes) != str:
                            has_note = True
                            break
                    for notes in notes_in_beat:
                        if has_note is True and type(notes) != str and len(notes) != 0:
                            merge_cell_nums.append(
                                [(now_row, now_column), (now_row + num_rows - 1, now_column)]
                            )
                            bef_sound_in_measure = notes
                        elif (
                            has_note is True
                            and bef_has_note is True
                            and notes == "-"
                            and bef_sound_in_measure is not None
                        ):
                            merge_cell_nums[-1][1] = (
                                now_row + num_rows - 1,
                                now_column + cells_per_notes - 1,
                            )
                        elif notes == "r":
                            bef_sound_in_measure = None

                        # 結合セル
                        for _row in range(now_row, now_row + num_rows):
                            self._merge_cells(
                                _row,
                                now_column,
                                _row,
                                now_column + cells_per_notes - 1,
                            )
                        width_dict = self._get_cell_widths(
                            self.ws, now_column, now_column + cells_per_notes - 1
                        )
                        shorten = self.shorten
                        # shorten = (
                        #     True
                        #     if sum(width_dict.values()) < self.shorten_width
                        #     else False
                        # )
                        # 音をプロット
                        bef_sound = self._plot_sound(
                            now_row, now_column, notes, bef_sound, num_rows, shorten
                        )
                        now_column += cells_per_notes
                    bef_has_note = has_note
                    beat_idx += 1
                # 伸ばすセルは結合（拍を超えていても結合、小節を挟んだ場合は結合しない）
                if len(merge_cell_nums) > 1:
                    for (_row1, _col1), (_row2, _col2) in merge_cell_nums:
                        for _row in range(_row1, _row2 + 1):
                            self._merge_cells(_row, _col1, _row, _col2)
            now_row += num_rows

            if progress_bar:
                progress_bar_value += progress_tick
                progress_bar["value"] = int(progress_bar_value)
                progress_bar.update()

        # 奏者、楽器を書く
        self._plot_player_and_instrument(
            start_row,
            now_row,
            start_column,
            player_idx=common_data.get_player_idx(),
            instrument_name=common_data.get_program_str(),
            difficulty=common_data.get_difficulty(self.tempo)
        )
        return now_row

    def _plot_beat_border(
        self, start_row, start_column, end_row, end_column, thick_flag
    ):
        for row in range(start_row, end_row + 1):
            start_cell_str = self._convert_cell_num(row, start_column)
            end_cell_str = self._convert_cell_num(row, end_column)
            print(start_row, start_column, "{}:{}".format(start_cell_str, end_cell_str))
            for row_cells in self.ws["{}:{}".format(start_cell_str, end_cell_str)]:
                for idx, cell in enumerate(row_cells):
                    border = Border(
                        top=self.non_border_side,
                        left=self.non_border_side,
                        bottom=self.non_border_side,
                        right=self.non_border_side,
                    )
                    if row == start_row:
                        border.top = self.medium_side
                    else:
                        border.top = self.thin_side
                    if row == end_row:
                        border.bottom = self.medium_side
                    else:
                        border.bottom = self.thin_side
                    if idx == 0 and thick_flag[0] == 1:
                        border.left = self.thin_side
                    elif idx == 0 and thick_flag[0] == 2:
                        border.left = self.medium_side
                    if idx == len(row_cells) - 1 and thick_flag[1] == 1:
                        border.right = self.thin_side
                    elif idx == len(row_cells) - 1 and thick_flag[1] == 2:
                        border.right = self.medium_side
                    cell.border = border

    def _plot_sound(self, row, column, notes, bef_sound, num_rows=1, shorten=False):
        # まずは休符を書く
        for _row in range(row, row + num_rows):
            self._plot_cell(_row, column, fill=self.rest_color_fill)

        if notes == "r":
            bef_sound = [None, None]
        elif notes == "-":
            if bef_sound[1] is not None:
                self._plot_cell(
                    row, column, fill=self.octave_color_fill[bef_sound[1]]
                )
        else:
            for num_note, note in enumerate(notes[::-1]):
                split_note = note.split("_")
                octave = 1 if len(split_note) == 1 else DICT_FOR_OCTAVE[split_note[1]]
                fill = self.octave_color_fill[octave]
                if shorten is False:
                    scale = DICT_FOR_CONVERT_GERMAN2JAPAN[split_note[0]]
                else:
                    scale = DICT_FOR_CONVERT_GERMAN2JAPAN_SHORTEN[split_note[0]]
                if num_note == 0:
                    bef_sound = [scale, octave]
                self._plot_cell(
                    row + num_note, column, scale, self.note_font, fill=fill)
        return bef_sound


def lcm(a, b):
    """
    ユークリッドの互除法による最小公倍数(least common multiple)算出
    注意: a >= b であること
    
    Parameters
    ----------
    a : int
        1つ目の数値
    b : int
        2つ目の数値
    
    Returns
    -------
    int
        aとbの最小公倍数
    """
    if a == 0:
        return 0
    elif b == 0:
        return a

    num_gcd = gcd(a, b)
    ret = (a * b) // num_gcd
    return ret


def gcd(a, b):
    """
    ユークリッドの互除法による最大公約数(gratest common divisor)算出
    注意: a >= b であること
    
    Parameters
    ----------
    a : int
        1つ目の数値
    b : int
        2つ目の数値
    
    Returns
    -------
    int
        aとbの最大公約数
    """
    if a == 0:
        return 0
    elif b == 0:
        return a

    a %= b
    while a != 0:
        a, b = b, a
        a %= b

    return b
