# coding: utf-8
from os.path import isfile
from openpyxl import load_workbook
from more_itertools import chunked
from colorsys import rgb_to_hsv

from .base import XlsxIOBase
from ._static_data import (
    DICT_FOR_NAME_CONVERT,
)

class XlsxLoader(XlsxIOBase):
    def __init__(self, filename, max_beat_num=8):
        """        
        Parameters
        ----------
        filename : str
            .xlsxのファイル名
        
        Raises
        ------
        OSError
            ファイルオープンエラー
        """
        if isfile(filename):
            self.wb = load_workbook(filename)
        else:
            raise OSError
        self.max_beat_num = max_beat_num

    def get_sound_list(self, sheet_name, start_col_char, end_col_char, row_list, get_rate=True):
        """
        共通音リスト(*), 共通音レートリスト(**)を返す関数

        *共通音リスト: 
            4/4 |♪♪|♪♪|♪♪|♪♪| --> [["c", "c"], ["c", "c"], ["c", "c"], ["c", "c"]]
                 ↑「ド」の音
        **共通音レートリスト:
            4/4 |♪♪|♪♪|♪♪|♪♪| --> ([1, 1, 1, 1], [[1, 1], [1, 1], [1, 1], [1, 1]])
                小節ごとの長さ関係・小節内の音の長さ関係を示したもの
        
        Parameters
        ----------
        sheet_name : str
            シート名
        start_col_char : char
            列名（英語）
        end_col_char : char
            列名（英語）
        row_list : list of int
            行リスト
        get_rate : bool
            横幅のレートを得るかどうか
        
        Returns
        -------
        list of list of list of str
            共通音リスト
        list of tuple
            共通音レートリスト
        """
        if sheet_name not in self.get_sheetnames():
            print("[ERROR] Invalid sheet name: {}".format(sheet_name))
            return None
        sheet = self.wb[sheet_name]
        start_col = self._convert_column_str(start_col_char.upper())
        end_col = self._convert_column_str(end_col_char.upper())
        merged_cells = self._get_merged_cells(sheet)
        cell_width_dict = self._get_cell_widths(sheet, start_col, end_col)

        all_notes = []
        all_rates = []
        measure_num = 1
        for row_num in row_list:
            beat_start_col = start_col
            while beat_start_col < end_col + 1:
                # [[ges_minus1, -], [-, -], [-, -], [r, r]] のような小節内の音リストを得る
                if (row_num, beat_start_col) in merged_cells:
                    print("skip(1)", row_num, beat_start_col)
                    beat_start_col += 1
                    continue
                val, color, left_border_style, right_border_style = self._get_cell_info(sheet, row_num, beat_start_col)
                note = self._get_note(val, color)
                if note is None:
                    beat_start_col += 1
                    continue
                notes = [[note]]
                cells = [[(row_num, beat_start_col)]]
                beat_num = 0
                bef_right_border_style = None
                bef_beat_start_col = beat_start_col
                for col_num in range(beat_start_col + 1, beat_start_col + 64):
                    val, color, left_border_style, right_border_style = self._get_cell_info(sheet, row_num, col_num)
                    if (row_num, col_num) in merged_cells:
                        bef_right_border_style = right_border_style
                        cells[beat_num].append((row_num, col_num))
                        continue
                    if (
                        (bef_right_border_style == "medium" or bef_right_border_style == "thick") or \
                        (left_border_style == "medium" or left_border_style == "thick")
                    ) or (col_num > end_col):
                        beat_start_col = col_num
                        break
                    elif (
                        (bef_right_border_style == "thin" or bef_right_border_style == "hair") or \
                        (left_border_style == "thin" or left_border_style == "hair")
                    ):
                        beat_num += 1
                        if beat_num >= self.max_beat_num:
                            beat_start_col = col_num
                            print("LIMIT BEAT", col_num)
                            break
                        else:
                            notes.append([])
                            cells.append([])
                    note = self._get_note(val, color)
                    if note is not None:
                        notes[beat_num].append(note)
                        cells[beat_num].append((row_num, col_num))
                    bef_right_border_style = right_border_style
                else:
                    print("ERROR: No border in next 64 cells")

                # widthからbeatの割合を算出
                widths_in_measure = []
                width_rates_in_beats = []
                widths_in_beats = []
                for cells_in_beat in cells:
                    widths_in_beat = []
                    sum_widths_in_beat = 0
                    for _, col_num in cells_in_beat:
                        width = cell_width_dict[col_num]
                        sum_widths_in_beat += width
                        if (row_num, col_num) not in merged_cells:
                            widths_in_beat.append(width)
                        else:
                            widths_in_beat[-1] += width
                    min_width = min(widths_in_beat)
                    widths_in_beats.append(widths_in_beat)
                    width_rates_in_beats.append([round(width / min_width) for width in widths_in_beat])
                    widths_in_measure.append(sum_widths_in_beat) 
                min_width = min(widths_in_measure)
                width_rates_in_measure = [round(width / min_width) for width in widths_in_measure]
                print("{0} [{1}{3}:{2}{3}]:".format(measure_num,
                    self._convert_column_num(bef_beat_start_col),
                    self._convert_column_num(beat_start_col), row_num
                ))
                print("    w_rate(beats):", widths_in_measure)
                print("  w_rate(in beat):", widths_in_beats)
                print("            notes:", notes)
                all_rates.append((width_rates_in_measure, width_rates_in_beats))
                all_notes.append(notes)
                measure_num += 1
        if get_rate:
            return all_notes, all_rates
        else:
            return all_notes, None

    def _convert_to_german_note(self, value):
        """
        セル内の文字をドイツ音名に変換する関数
        
        Parameters
        ----------
        value : str
            音の文字列
        
        Returns
        -------
        list of str
            ドイツ音名の音リスト
        """
        ret = []
        now_idx = 0
        eng_note = ""
        while now_idx < len(value):
            v = value[now_idx]
            add_note = False
            for candidates, (note_name, register_note) in DICT_FOR_NAME_CONVERT.items():
                for c in candidates:
                    if type(c) is str:  # 比較候補が1文字の場合
                        if v == c:
                            add_note = True
                            break
                    elif type(c) is list or type(c) is tuple:  # 比較候補が2文字以上の場合
                        if self._check_same_strings_or_not(value, now_idx, c) is True:
                            add_note = True
                            break
                if add_note:
                    if register_note is True and len(eng_note) != 0:
                        ret.append(eng_note)
                        eng_note = ""
                    if (eng_note == "E" or eng_note == "A") and note_name == "es":
                        eng_note += "s"
                    elif eng_note == "H" and note_name == "es":
                        eng_note = "B"
                    else:
                        eng_note += note_name
                    break
            now_idx += 1
        if len(eng_note) != 0:
            ret.append(eng_note)
        return ret

    def _get_note(self, value, color):
        """
        セルの内容を音に変換して返す関数
        
        Parameters
        ----------
        value : str
            セルの内容
        color : str
            背景色("red", "yellow", "green", "cyan", "blue", "purple", Noneのいずれか)
        
        Returns
        -------
        list of str or None
            セル内の音リスト
        """
        if color == "white":
            return None
        if color == "gray" or color is None:
            return ["r"]
        if value is None or value.strip() == "":
            return ["-"]
        ret = self._convert_to_german_note(value.strip())
        if len(ret) == 0:
            return None
        if color == "cyan" or color == "blue":
            for i, r in enumerate(ret):
                if r == "r" or r == "-":
                    continue
                ret[i] += "_minus1"
        elif color == "yellow":
            pass
        elif color == "red":
            for i, r in enumerate(ret):
                if r == "r" or r == "-":
                    continue
                ret[i] += "_plus1"
        elif color == "purple":
            for i, r in enumerate(ret):
                if r == "r" or r == "-":
                    continue
                ret[i] += "_plus2"
        return ret

    def _get_cell_info(self, sheet, row_num, col_num):
        """
        セルの情報を返す関数
        
        Parameters
        ----------
        sheet : openpyxl.WorkSheet
            エクセルのワークシート(openpyxl形式)
        row_num : int
            行番号
        col_num : int
            列番号
        
        Returns
        -------
        str
            セルの値
        str or None
            セルの色
        str
            左罫線のスタイル
            Value must be one of {‘double’, ‘dashed’, ‘thin’, ‘medium’, ‘mediumDashDot’,
            ‘dashDot’, ‘thick’, ‘mediumDashed’, ‘hair’, ‘dotted’, ‘slantDashDot’,
            ‘mediumDashDotDot’, ‘dashDotDot’}
        str
            右罫線のスタイル
            Value must be one of {‘double’, ‘dashed’, ‘thin’, ‘medium’, ‘mediumDashDot’,
            ‘dashDot’, ‘thick’, ‘mediumDashed’, ‘hair’, ‘dotted’, ‘slantDashDot’,
            ‘mediumDashDotDot’, ‘dashDotDot’}
        """
        cell = sheet.cell(row=row_num, column=col_num)
        rgb = [int(c1 + c2, 16) for c1, c2 in chunked(cell.fill.fgColor.rgb[2:], 2)]
        hsv = rgb_to_hsv(*rgb)
        if hsv[1] > .01:
            color = self.judge_color(hsv[0])
        elif hsv[2] > 0:
            color = "gray"
        else:
            color = "white"
        # print(row_num, col_num, cell.value, cell.fill.fgColor.rgb, cell.fill.bgColor.rgb, hsv, color, cell.border.left.style, cell.border.right.style)
        return cell.value, color, cell.border.left.style, cell.border.right.style

    def _get_merged_cells(self, sheet):
        """
        結合セルの座標リストを返す関数
        
        Parameters
        ----------
        sheet : openpyxl.WorkSheet
            エクセルのワークシート(openpyxl形式)
        
        Returns
        -------
        list of tuple of int
            (row, col)の数値が入ったリスト
        """
        merged_cells = []
        for merged_cell_str in sheet.merged_cell_ranges:
            start_cell, end_cell = merged_cell_str.split(":")
            merge_start_row, merge_start_col = self._convert_cell_str(start_cell)
            merge_end_row, merge_end_col = self._convert_cell_str(end_cell)
            if merge_start_row != merge_end_row:
                continue  # 2行以上にわたっている結合セルなので、楽譜以外の結合セルと判定
            else:
                for i in range(merge_start_col + 1, merge_end_col + 1):  # 結合最初のセルは範囲外 かつ range関数は最後の番号未満までなのでお互い+1
                    merged_cells.append((merge_start_row, i))
        return merged_cells

    def _check_same_strings_or_not(self, value, now_idx, candidate):
        """
        セルの値と音の文字候補が一致しているかどうかを返す関数
        
        Parameters
        ----------
        value : str
            セルの値
        now_idx : int
            現在の文字数
        candidate : 音の文字候補
            音の文字候補
        
        Returns
        -------
        bool
            セルの値と音の文字候補が合っているかどうか
        """
        is_same = True
        for i, _c in enumerate(candidate):
            if value[now_idx + i] != _c:
                is_same = False
                break
        return is_same
